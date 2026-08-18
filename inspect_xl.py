import openpyxl
wb = openpyxl.load_workbook('SD Checklists.xlsx')

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'\n{"="*60}')
    print(f'SHEET: {sheet}  ({ws.max_row} rows x {ws.max_column} cols)')
    print(f'{"="*60}')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or any(v is not None for v in row):
            print(f'  {i+1}: {row}')
        if i > 200:
            print('  ... (truncated)')
            break
