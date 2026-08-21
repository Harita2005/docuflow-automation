import openpyxl
from pathlib import Path

excel_path = Path("SD SCHEMA AND WORKFLOW DETAILS.xlsx")
wb = openpyxl.load_workbook(str(excel_path), data_only=True)

print("Workbook sheets:", wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    print(f"\nSheet: {s} | Rows: {ws.max_row} | Cols: {ws.max_column}")
    print("Headers:", headers)
    # Print first 3 rows
    for r in range(2, min(5, ws.max_row + 1)):
        row_vals = [ws.cell(r, col).value for col in range(1, ws.max_column + 1)]
        print(f"  Row {r}: {row_vals}")
