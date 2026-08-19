import os
import json
import openpyxl
from pathlib import Path
from sqlalchemy.orm import Session
from app.database import SessionLocal, engine, Base
from app.models import ChecklistRule

def seed_unified_rules():
    print("==================================================================")
    print(">>> SEEDING UNIFIED CHECKLIST RULES FROM EXCEL")
    print("==================================================================")
    
    from sqlalchemy import text
    with engine.connect() as connection:
        try:
            connection.execute(text("IF OBJECT_ID('checklist_rules', 'U') IS NOT NULL DROP TABLE checklist_rules;"))
            connection.commit()
        except Exception:
            pass
            
    # Ensure tables are created
    Base.metadata.create_all(bind=engine)
    db: Session = SessionLocal()
    
    try:
        # Clear old rules
        db.query(ChecklistRule).delete()
        db.commit()
        
        BASE_DIR = Path(__file__).resolve().parent
        excel_path = BASE_DIR.parent / "SD Checklists.xlsx"
        wb = openpyxl.load_workbook(str(excel_path))
        
        # ----------------------------------------------------
        # 1. SEED SHEET 4: Workflow-Based DefaultChecklist
        # ----------------------------------------------------
        ws_wf = wb['Workflow-Based DefaultChecklist']
        wf_count = 0
        for r in range(2, ws_wf.max_row + 1):
            wf_name = ws_wf.cell(row=r, column=1).value
            stage_name = ws_wf.cell(row=r, column=2).value
            chk_str = ws_wf.cell(row=r, column=3).value
            
            if not wf_name or not stage_name or not chk_str:
                continue
                
            # Clean and split checklist items
            raw_items = chk_str.split(',')
            cleaned_items = []
            temp = ""
            for item in raw_items:
                candidate = item.strip()
                if not candidate:
                    continue
                if "(" in candidate and ")" not in candidate:
                    temp = candidate
                elif temp:
                    temp += ", " + candidate
                    if ")" in candidate:
                        cleaned_items.append(temp)
                        temp = ""
                else:
                    cleaned_items.append(candidate)
                    
            final_items = [x.strip().replace("\u00a0", " ").strip(", ") for x in cleaned_items if x.strip()]
            
            for seq, item_text in enumerate(final_items, start=1):
                rule = ChecklistRule(
                    rule_name=f"Workflow Fallback: {wf_name}",
                    workflow_profile=wf_name,
                    stage_name=stage_name,
                    item_text=item_text,
                    is_mandatory=True,
                    is_active=True,
                    sequence_order=seq
                )
                db.add(rule)
                wf_count += 1
            
        print(f"[OK] Seeded {wf_count} Workflow Fallback Checklist Items.")
        
        # ----------------------------------------------------
        # 2. SEED SHEET 2: Checklist configured Category
        # ----------------------------------------------------
        ws_cat = wb['Checklist configured Category']
        cat_count = 0
        for r in range(2, ws_cat.max_row + 1):
            item_text = ws_cat.cell(row=r, column=1).value
            stage_name = ws_cat.cell(row=r, column=2).value
            company = ws_cat.cell(row=r, column=5).value
            category = ws_cat.cell(row=r, column=6).value
            branch = ws_cat.cell(row=r, column=8).value
            
            if not stage_name or not company or not category or not item_text:
                continue
                
            rule = ChecklistRule(
                rule_name=f"Category Rule: {company.strip()} - {category.strip()}",
                division=company.strip() if company.strip() != "ALL" else None,
                category=category.strip() if category.strip() != "ALL" else None,
                branch=str(branch).strip() if str(branch).strip() not in ("ALL", "None", "") else None,
                stage_name=stage_name.strip(),
                item_text=item_text.strip().replace("\u00a0", " "),
                is_mandatory=True,
                is_active=True,
                sequence_order=r - 1
            )
            db.add(rule)
            cat_count += 1
            
        print(f"[OK] Seeded {cat_count} Category Specific Checklist Items.")
        
        # ----------------------------------------------------
        # 3. SEED SHEET 3: VCC ChecklistConfiguredCategory
        # ----------------------------------------------------
        ws_vcc = wb['VCC ChecklistConfiguredCategory']
        vcc_items = [
            "Bill Amount Verified",
            "Bill No ,Date & Address Verified",
            "Documents Attached",
            "Debit/Credit Note Verified",
            "Gofrugal Entry with narration Verified",
            "PO Verified",
            "Showroom Inward details",
            "Tax portion verified (GST, TDS, etc..)",
            "Vendor GST no,  Signaure Verified",
            "Vendor Name",
            "Showroom Name Verified"
        ]
        
        vcc_groups = set()
        for r in range(2, ws_vcc.max_row + 1):
            stage_name = ws_vcc.cell(row=r, column=2).value
            category = ws_vcc.cell(row=r, column=6).value
            branch = ws_vcc.cell(row=r, column=8).value
            
            if not stage_name or not category:
                continue
            vcc_groups.add((category.strip(), str(branch).strip(), stage_name.strip()))
            
        vcc_count = 0
        for category, branch, stage_name in vcc_groups:
            for seq, item_text in enumerate(vcc_items, start=1):
                rule = ChecklistRule(
                    rule_name=f"VCC Category Rule: {category}",
                    division="VCC",
                    category=category if category != "ALL" else None,
                    branch=branch if branch not in ("ALL", "None", "", "None") else None,
                    stage_name=stage_name,
                    item_text=item_text,
                    is_mandatory=True,
                    is_active=True,
                    sequence_order=seq
                )
                db.add(rule)
                vcc_count += 1
            
        print(f"[OK] Seeded {vcc_count} VCC Division Checklist Items.")
        
        db.commit()
        print(f"\n[SUCCESS] Successfully seeded {db.query(ChecklistRule).count()} total unified checklist rules!")
        
    except Exception as e:
        db.rollback()
        print(f"[ERROR] Seeding failed: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    seed_unified_rules()
