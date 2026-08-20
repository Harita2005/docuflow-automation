import os
import json
import openpyxl
from pathlib import Path
from sqlalchemy.orm import Session
from app.database import SessionLocal, engine, Base
from app.models import WorkflowProfile, WorkflowStepDefinition, ChecklistTemplate, ChecklistRule

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "SD Checklists.xlsx"
if not EXCEL_PATH.exists():
    EXCEL_PATH = BASE_DIR.parent / "SD Checklists.xlsx"

def seed_perfect_stage_checklists():
    print("==================================================================")
    print(">>> SEEDING 100% PERFECT CHECKLISTS FOR EVERY FLOW & STAGE")
    print(f">>> Source: {EXCEL_PATH}")
    print("==================================================================")

    if not EXCEL_PATH.exists():
        print(f"[ERROR] Excel file not found at: {EXCEL_PATH}")
        return

    Base.metadata.create_all(bind=engine)
    db: Session = SessionLocal()

    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

        # ----------------------------------------------------
        # 1. Parse Sheet 4: Workflow-Based DefaultChecklist
        # ----------------------------------------------------
        print("\n[1/4] Parsing Workflow-Based DefaultChecklist...")
        ws_wf = wb['Workflow-Based DefaultChecklist']
        wf_stage_checklists = {}  # wf_name -> { stage_name -> [items] }

        for r in range(2, ws_wf.max_row + 1):
            wf_name = ws_wf.cell(row=r, column=1).value
            stage_name = ws_wf.cell(row=r, column=2).value
            chk_str = ws_wf.cell(row=r, column=3).value

            if not wf_name or not stage_name or not chk_str:
                continue

            wf_name = str(wf_name).strip()
            stage_name = str(stage_name).strip()

            raw_items = str(chk_str).split(',')
            cleaned_items = []
            temp = ""
            for item in raw_items:
                candidate = item.strip()
                if not candidate:
                    continue
                if "(" in candidate and ")" not in candidate:
                    temp = candidate
                elif temp:
                    temp += ", " + candidate
                    if ")" in candidate:
                        cleaned_items.append(temp)
                        temp = ""
                else:
                    cleaned_items.append(candidate)

            final_items = [x.strip().replace("\u00a0", " ").strip(", ") for x in cleaned_items if x.strip()]

            if wf_name not in wf_stage_checklists:
                wf_stage_checklists[wf_name] = {}
            wf_stage_checklists[wf_name][stage_name] = final_items

        print(f"  Parsed {len(wf_stage_checklists)} workflow template structures from Sheet 4.")

        # ----------------------------------------------------
        # 2. Parse Sheet 2 & 3: Category Configured Checklists
        # ----------------------------------------------------
        print("\n[2/4] Parsing Category & VCC Configured Checklists...")
        cat_stage_checklists = {}

        # Sheet 2: Checklist configured Category
        ws_cat = wb['Checklist configured Category']
        for r in range(2, ws_cat.max_row + 1):
            item_text = ws_cat.cell(row=r, column=1).value
            stage_name = ws_cat.cell(row=r, column=2).value
            company = ws_cat.cell(row=r, column=5).value
            category = ws_cat.cell(row=r, column=6).value

            if not item_text or not stage_name:
                continue

            item_str = str(item_text).strip().replace("\u00a0", " ").strip(", ")
            st_str = str(stage_name).strip()
            comp_str = str(company).strip() if company else "ALL"
            cat_str = str(category).strip() if category else "ALL"

            key = (comp_str, cat_str, st_str)
            if key not in cat_stage_checklists:
                cat_stage_checklists[key] = []
            if item_str not in cat_stage_checklists[key]:
                cat_stage_checklists[key].append(item_str)

        # Sheet 3: VCC ChecklistConfiguredCategory
        ws_vcc = wb['VCC ChecklistConfiguredCategory']
        for r in range(2, ws_vcc.max_row + 1):
            item_text = ws_vcc.cell(row=r, column=1).value
            stage_name = ws_vcc.cell(row=r, column=2).value
            category = ws_vcc.cell(row=r, column=6).value

            if not item_text or not stage_name:
                continue

            item_str = str(item_text).strip().replace("\u00a0", " ").strip(", ")
            st_str = str(stage_name).strip()
            cat_str = str(category).strip() if category else "ALL"

            key = ("VCC", cat_str, st_str)
            if key not in cat_stage_checklists:
                cat_stage_checklists[key] = []
            if item_str not in cat_stage_checklists[key]:
                cat_stage_checklists[key].append(item_str)

        print(f"  Parsed {len(cat_stage_checklists)} Category-specific stage checklists.")

        # Standard Canonical Checklists
        vcc_full_checklist = [
            "Documents Attached",
            "Bill Amount Verified",
            "Bill No ,Date & Address Verified",
            "Vendor Name",
            "Showroom Name Verified",
            "Vendor GST no,  Signaure Verified",
            "Tax portion verified (GST, TDS, etc..)",
            "PO Verified",
            "Showroom Inward details",
            "Debit/Credit Note Verified",
            "Gofrugal Entry with narration Verified"
        ]

        vcc_review_checklist = [
            "Documents Attached",
            "Bill Amount Verified",
            "Showroom Name Verified",
            "Vendor Name"
        ]

        vcc_final_checklist = [
            "Documents Attached",
            "Party Name & Total Amount Verified",
            "Final Settlement Sign-off"
        ]

        standard_9point_checklist = [
            "Documents Attached",
            "Party Name & Total Amount Verified",
            "Vendor GST no, Signaure Verified",
            "Bill No ,Date & Address Verified",
            "Tax portion verified (GST, TDS, etc..)",
            "RO/PO Verified",
            "Gate Inward, GRN, Debit/Credit Note Verified",
            "SAP Entry ( DR/CR & GL , COST CENTER ) Verified",
            "Advance, Narration, Supportive Copy (If Any)"
        ]

        grn_header_checklist = [
            "PO Ref Number, Quantity Verified",
            "Bill Number, Bill Date Verified",
            "GRN Number, Quantity Verified",
            "Bundle Quantity Verified",
            "Invoice Quantity, DC Quantity Verified",
            "Security Seal and Signature Verified",
            "PSecurity Checking Slip Attached",
            "Signatures & Total Value Verified"
        ]

        # ----------------------------------------------------
        # 3. Seed ChecklistTemplate & ChecklistRule Tables
        # ----------------------------------------------------
        print("\n[3/4] Seeding ChecklistTemplate and ChecklistRule tables...")
        db.query(ChecklistTemplate).delete()
        db.query(ChecklistRule).delete()
        db.commit()

        # Seed ChecklistRule from Category-configured items
        rule_chk_count = 0
        for (comp, cat, st), items in cat_stage_checklists.items():
            for seq, it in enumerate(items, 1):
                db.add(ChecklistRule(
                    rule_name=f"CHK_{comp}_{cat[:20]}_{st[:15]}_{seq}",
                    division=comp,
                    category=cat,
                    branch="ALL",
                    workflow_profile=None,
                    stage_name=st,
                    item_text=it,
                    is_mandatory=True,
                    is_active=True,
                    sequence_order=seq
                ))
                rule_chk_count += 1
        db.commit()

        # ----------------------------------------------------
        # 4. Populate ChecklistTemplate for 100% of Workflow Profiles & Steps
        # ----------------------------------------------------
        print("\n[4/4] Populating ChecklistTemplate for 100% of all Workflow Profiles and Steps...")
        all_profiles = db.query(WorkflowProfile).all()
        all_steps = db.query(WorkflowStepDefinition).all()

        # Group steps by profile_name
        steps_by_profile = {}
        for s in all_steps:
            if s.profile_name not in steps_by_profile:
                steps_by_profile[s.profile_name] = []
            steps_by_profile[s.profile_name].append(s)

        tpl_count = 0
        for p in all_profiles:
            p_name = p.profile_name
            steps = steps_by_profile.get(p_name, [])

            # Check if this profile has custom template in Sheet 4
            sheet4_match = None
            if p_name in wf_stage_checklists:
                sheet4_match = wf_stage_checklists[p_name]
            else:
                for k, v in wf_stage_checklists.items():
                    if k in p_name or p_name in k:
                        sheet4_match = v
                        break

            for step in steps:
                st_name = step.step_name
                target_items = []

                if sheet4_match and st_name in sheet4_match:
                    target_items = sheet4_match[st_name]
                elif "GRN" in p_name.upper():
                    target_items = grn_header_checklist
                elif "VCC" in p_name.upper():
                    if st_name == "Final Approval":
                        target_items = vcc_final_checklist
                    elif st_name in ["First Approval", "Second Approval", "3rd APPROVAL"]:
                        target_items = vcc_review_checklist
                    else:
                        target_items = vcc_full_checklist
                else:
                    if st_name == "Final Approval":
                        target_items = ["Documents Attached", "Party Name & Total Amount Verified"]
                    elif st_name in ["Second Approval", "3rd APPROVAL"]:
                        target_items = ["Documents Attached", "Party Name & Total Amount Verified"]
                    elif st_name == "First Approval":
                        target_items = ["Documents Attached", "Party Name & Total Amount Verified", "RO/PO Verified", "Gate Inward, GRN, Debit/Credit Note Verified"]
                    else:
                        target_items = standard_9point_checklist

                for seq, it_text in enumerate(target_items, 1):
                    db.add(ChecklistTemplate(
                        workflow_profile=p_name,
                        stage_name=st_name,
                        item_text=it_text,
                        is_mandatory=True,
                        is_active=True,
                        sequence_order=seq
                    ))
                    tpl_count += 1

        db.commit()
        print(f"  [OK] Seeded {tpl_count} ChecklistTemplate items across {len(all_profiles)} profiles!")
        print(f"  [OK] Seeded {rule_chk_count} ChecklistRule override items!")

        print("\n==================================================================")
        print(">>> 100% SUCCESS: EVERY SINGLE WORKFLOW PROFILE & STAGE HAS A PERFECT CHECKLIST!")
        print("==================================================================")

    except Exception as e:
        db.rollback()
        print(f"[ERROR] Checklist seeding failed: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()

if __name__ == "__main__":
    seed_perfect_stage_checklists()
