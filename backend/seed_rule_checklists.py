import os
import json
import openpyxl
from pathlib import Path
from sqlalchemy.orm import Session
from app.database import SessionLocal, engine, Base
from app.models import BusinessRule, WorkflowProfile, WorkflowStepDefinition

def seed_rule_checklists():
    print("==================================================================")
    print(">>> SEEDING CHECKLISTS DIRECTLY INTO BUSINESS RULES")
    print("==================================================================")
    
    db: Session = SessionLocal()
    
    try:
        BASE_DIR = Path(__file__).resolve().parent
        excel_path = BASE_DIR.parent / "SD Checklists.xlsx"
        
        if not excel_path.exists():
            print(f"[ERROR] Excel file not found at {excel_path}")
            return
            
        wb = openpyxl.load_workbook(str(excel_path))
        
        # ----------------------------------------------------
        # 1. Parse Sheet 4: Workflow-Based DefaultChecklist
        # ----------------------------------------------------
        print("Parsing Workflow-Based DefaultChecklist...")
        ws_wf = wb['Workflow-Based DefaultChecklist']
        wf_checklists = {} # wf_name -> { stage_name -> [items] }
        
        for r in range(2, ws_wf.max_row + 1):
            wf_name = ws_wf.cell(row=r, column=1).value
            stage_name = ws_wf.cell(row=r, column=2).value
            chk_str = ws_wf.cell(row=r, column=3).value
            
            if not wf_name or not stage_name or not chk_str:
                continue
                
            wf_name = wf_name.strip()
            stage_name = stage_name.strip()
            
            # Clean items
            raw_items = chk_str.split(',')
            cleaned_items = []
            temp = ""
            for item in raw_items:
                candidate = item.strip()
                if not candidate:
                    continue
                if "(" in candidate and ")" not in candidate:
                    temp = candidate
                elif temp:
                    temp += ", " + candidate
                    if ")" in candidate:
                        cleaned_items.append(temp)
                        temp = ""
                else:
                    cleaned_items.append(candidate)
                    
            final_items = [x.strip().replace("\u00a0", " ").strip(", ") for x in cleaned_items if x.strip()]
            
            if wf_name not in wf_checklists:
                wf_checklists[wf_name] = {}
            wf_checklists[wf_name][stage_name] = final_items
            
        # ----------------------------------------------------
        # 2. Parse Sheet 2: Checklist configured Category
        # ----------------------------------------------------
        print("Parsing Checklist configured Category...")
        ws_cat = wb['Checklist configured Category']
        cat_checklists = {} # CategoryName -> { stage_name -> [items] }
        
        for r in range(2, ws_cat.max_row + 1):
            item_text = ws_cat.cell(row=r, column=1).value
            stage_name = ws_cat.cell(row=r, column=2).value
            category_name = ws_cat.cell(row=r, column=4).value
            
            if not stage_name or not category_name or not item_text:
                continue
                
            cat_name = category_name.strip()
            stage_name = stage_name.strip()
            
            if cat_name not in cat_checklists:
                cat_checklists[cat_name] = {}
            if stage_name not in cat_checklists[cat_name]:
                cat_checklists[cat_name][stage_name] = []
                
            cat_checklists[cat_name][stage_name].append(item_text.strip().replace("\u00a0", " "))
            
        # ----------------------------------------------------
        # 3. Match and Update Business Rules
        # ----------------------------------------------------
        print("Updating Business Rules with Checklist data...")
        rules = db.query(BusinessRule).all()
        updated_count = 0
        
        # Standard VCC items for fallback
        vcc_items = [
            "Bill Amount Verified",
            "Bill No ,Date & Address Verified",
            "Documents Attached",
            "Debit/Credit Note Verified",
            "Gofrugal Entry with narration Verified",
            "PO Verified",
            "Showroom Inward details",
            "Tax portion verified (GST, TDS, etc..)",
            "Vendor GST no,  Signaure Verified",
            "Vendor Name",
            "Showroom Name Verified"
        ]
        
        for r in rules:
            target_checklists = {}
            wf_id = r.target_workflow_id
            
            # Match by workflow profile id (e.g. ACM_GRN_Header_2stages)
            if wf_id in wf_checklists:
                target_checklists.update(wf_checklists[wf_id])
                
            # Match by category profile (e.g. ACM_GRN_HEADER)
            # Find any category name that contains or matches
            for cat_name, stages in cat_checklists.items():
                if cat_name in wf_id or wf_id in cat_name:
                    target_checklists.update(stages)
                    
            # Fallback for VCC rules to VCC default checklists
            if "VCC" in wf_id or (r.conditions_json and "VCC" in r.conditions_json):
                # Retrieve all steps for this VCC workflow profile
                steps = db.query(WorkflowStepDefinition).filter(
                    WorkflowStepDefinition.profile_name == wf_id
                ).all()
                for step in steps:
                    if step.step_name not in target_checklists:
                        # Match sheet 4 if present
                        vcc_wf_mapped = None
                        for key in wf_checklists:
                            if key in wf_id or wf_id in key:
                                vcc_wf_mapped = key
                                break
                        if vcc_wf_mapped and step.step_name in wf_checklists[vcc_wf_mapped]:
                            target_checklists[step.step_name] = wf_checklists[vcc_wf_mapped][step.step_name]
                        else:
                            target_checklists[step.step_name] = vcc_items
                            
            if target_checklists:
                r.stage_checklists = json.dumps(target_checklists)
                updated_count += 1
                
        db.commit()
        print(f"\n[SUCCESS] Successfully mapped and saved checklists onto {updated_count} Business Rules!")
        
    except Exception as e:
        db.rollback()
        print(f"[ERROR] Seeding failed: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    seed_rule_checklists()
